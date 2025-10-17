"""
Data Export Utilities

Provides functions to export generated data in various formats.
"""

import json
import csv
import io
import zipfile
from typing import Any, Dict, List, Optional, Union
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class DataExporter:
    """Handles data export in various formats."""
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "synthetic_data.csv") -> str:
        """Export data to CSV format."""
        if not data:
            return ""
        
        output = io.StringIO()
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue()
    
    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], format_type: str = "array") -> str:
        """Export data to JSON format."""
        if not data:
            return "[]"
        
        if format_type == "array":
            return json.dumps(data, indent=2, default=str)
        elif format_type == "lines":
            # Line-delimited JSON
            lines = []
            for record in data:
                lines.append(json.dumps(record, default=str))
            return "\n".join(lines)
        else:
            return json.dumps(data, indent=2, default=str)
    
    @staticmethod
    def export_to_parquet(data: List[Dict[str, Any]]) -> bytes:
        """Export data to Parquet format."""
        if not data:
            return b""
        
        df = pd.DataFrame(data)
        table = pa.Table.from_pandas(df)
        buffer = io.BytesIO()
        pq.write_table(table, buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = "synthetic_data.xlsx") -> bytes:
        """Export data to Excel format."""
        if not data:
            return b""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthetic Data"
        
        # Add headers
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, record in enumerate(data, 2):
            for col, (key, value) in enumerate(record.items(), 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_sql(data: List[Dict[str, Any]], table_name: str = "synthetic_data") -> str:
        """Export data as SQL INSERT statements."""
        if not data:
            return ""
        
        sql_statements = []
        
        # Get column names
        columns = list(data[0].keys())
        column_list = ", ".join([f"`{col}`" for col in columns])
        
        # Generate INSERT statements
        for record in data:
            values = []
            for col in columns:
                value = record.get(col)
                if value is None:
                    values.append("NULL")
                elif isinstance(value, str):
                    # Escape single quotes
                    escaped_value = value.replace("'", "''")
                    values.append(f"'{escaped_value}'")
                elif isinstance(value, (int, float)):
                    values.append(str(value))
                else:
                    values.append(f"'{str(value)}'")
            
            values_list = ", ".join(values)
            sql_statements.append(f"INSERT INTO `{table_name}` ({column_list}) VALUES ({values_list});")
        
        return "\n".join(sql_statements)
    
    @staticmethod
    def export_to_pandas_code(data: List[Dict[str, Any]], variable_name: str = "df") -> str:
        """Export data as Python Pandas DataFrame code."""
        if not data:
            return f"{variable_name} = pd.DataFrame()"
        
        # Convert data to a format that can be easily represented in code
        code_lines = [f"{variable_name} = pd.DataFrame(["]
        
        for i, record in enumerate(data):
            if i > 0:
                code_lines.append(",")
            
            # Format record as dictionary
            record_str = "{"
            for j, (key, value) in enumerate(record.items()):
                if j > 0:
                    record_str += ", "
                
                if isinstance(value, str):
                    escaped_value = value.replace("'", "\\'")
                    record_str += f"'{key}': '{escaped_value}'"
                elif value is None:
                    record_str += f"'{key}': None"
                else:
                    record_str += f"'{key}': {value}"
            
            record_str += "}"
            code_lines.append(f"    {record_str}")
        
        code_lines.append("])")
        
        return "\n".join(code_lines)
    
    @staticmethod
    def create_zip_archive(files: Dict[str, Union[str, bytes]], 
                          archive_name: str = "synthetic_data.zip") -> bytes:
        """Create a ZIP archive containing multiple files."""
        buffer = io.BytesIO()
        
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, content in files.items():
                if isinstance(content, str):
                    zip_file.writestr(filename, content)
                else:
                    zip_file.writestr(filename, content)
        
        return buffer.getvalue()
    
    @staticmethod
    def export_with_compression(data: List[Dict[str, Any]], 
                              format_type: str = "csv",
                              compress: bool = False) -> Union[str, bytes]:
        """Export data with optional compression."""
        if format_type == "csv":
            content = DataExporter.export_to_csv(data)
            if compress:
                files = {"synthetic_data.csv": content}
                return DataExporter.create_zip_archive(files)
            return content
        
        elif format_type == "json":
            content = DataExporter.export_to_json(data)
            if compress:
                files = {"synthetic_data.json": content}
                return DataExporter.create_zip_archive(files)
            return content
        
        elif format_type == "parquet":
            return DataExporter.export_to_parquet(data)
        
        elif format_type == "excel":
            return DataExporter.export_to_excel(data)
        
        elif format_type == "sql":
            content = DataExporter.export_to_sql(data)
            if compress:
                files = {"synthetic_data.sql": content}
                return DataExporter.create_zip_archive(files)
            return content
        
        elif format_type == "pandas":
            content = DataExporter.export_to_pandas_code(data)
            if compress:
                files = {"synthetic_data.py": content}
                return DataExporter.create_zip_archive(files)
            return content
        
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    @staticmethod
    def get_export_info(data: List[Dict[str, Any]], format_type: str) -> Dict[str, Any]:
        """Get information about the export."""
        if not data:
            return {
                'format': format_type,
                'size_bytes': 0,
                'record_count': 0,
                'field_count': 0
            }
        
        # Get basic info
        record_count = len(data)
        field_count = len(data[0].keys()) if data else 0
        
        # Calculate size
        if format_type == "csv":
            content = DataExporter.export_to_csv(data)
            size_bytes = len(content.encode('utf-8'))
        elif format_type == "json":
            content = DataExporter.export_to_json(data)
            size_bytes = len(content.encode('utf-8'))
        elif format_type == "parquet":
            content = DataExporter.export_to_parquet(data)
            size_bytes = len(content)
        elif format_type == "excel":
            content = DataExporter.export_to_excel(data)
            size_bytes = len(content)
        elif format_type == "sql":
            content = DataExporter.export_to_sql(data)
            size_bytes = len(content.encode('utf-8'))
        elif format_type == "pandas":
            content = DataExporter.export_to_pandas_code(data)
            size_bytes = len(content.encode('utf-8'))
        else:
            size_bytes = 0
        
        return {
            'format': format_type,
            'size_bytes': size_bytes,
            'size_mb': round(size_bytes / (1024 * 1024), 2),
            'record_count': record_count,
            'field_count': field_count
        }
